"""Offline integration test for the --elumen-live build path.

SELF-CONTAINED + fully offline/deterministic. A minimal FakeClient (conftest's
URL-substring matcher) covers THREE live sources with OVERLAPPING courses so the
fetched eLumen prereq actually attaches to a catalog row the schedule produced:

  - schedule listing  -> ANATOMY 1, BIOLOGY 3, BIOLOGY 5 (the gated course AND
    both prereq targets are in the catalog);
  - Program Mapper    -> a program whose courses include ANATOMY 1 + BIOLOGY 3/5;
  - eLumen            -> the committed golden fixture
    (tests/fixtures/elumen_courses_LAMC_response.json) on "/public/courses".

The committed fixture's build_prereq_map output for ANATOMY 1 is the golden
"(BIOLOGY 3 OR BIOLOGY 5)". We assert that string lands in the catalog sheet,
that the coverage report is present, that the prerequisite_ordering detector is
ACTIVE and labelled REAL (not fixture-only), and that engine.run consumed the
workbook (its solver plan includes the gated course). A precedence test confirms
that supplying BOTH --elumen-live and --elumen-fixture uses LIVE and records a
warning.

NO NETWORK: every source is served by the injected FakeClient. The build path's
own httpx.Client is only opened when client is None, which never happens here, so
the test passes with no socket access (a third test pins that by failing on any
real socket connect).
"""
import json
import pathlib

import openpyxl
import pytest

import build_live_workbook
import engine
from tests.conftest import FakeClient

FIXTURES = pathlib.Path(__file__).parent / "fixtures"
ELUMEN_FIXTURE = str(FIXTURES / "elumen_courses_LAMC_response.json")

# --- minimal three-source route map (overlapping courses) ------------------
# The three courses overlap the eLumen golden fixture: the gated course
# (ANATOMY 1) plus both of its prereq targets (BIOLOGY 3, BIOLOGY 5), so the
# eLumen prereq "(BIOLOGY 3 OR BIOLOGY 5)" both ATTACHES to ANATOMY 1 AND finds
# its targets present in the catalog.

# Schedule listing shape: subjects[] -> courses[] (subject + catalogNbr) ->
# sections[] (each with meetings[]). sources.schedule builds course id as
# f"{subject} {catalogNbr}".
def _course(subject, catalog, title):
    return {
        "subject": subject, "catalogNbr": catalog, "descr": title, "units": "4",
        "sections": [{
            "classNbr": f"1{subject[:2]}{catalog}", "status": "Open",
            "classType": ["INP"],
            "meetings": [{"days": "MW", "times": "9-11", "room": "S1",
                          "instr": "Staff"}],
        }],
    }


_LISTING = {
    "subjects": [
        {"name": "ANATOMY", "courses": [_course("ANATOMY", "1", "Human Anatomy")]},
        {"name": "BIOLOGY", "courses": [
            _course("BIOLOGY", "3", "Introductory Biology"),
            _course("BIOLOGY", "5", "General Biology"),
        ]},
    ]
}

# Program Mapper chain shapes: home-page-content (programGroups) ->
# program-groups/<gid> (programs) -> programs/<pid> (pathways) ->
# program-maps/<mid> (pathwayElements with COURSE opportunities). The program's
# courses are exactly the three overlapping ones, so reconciliation matches all.
_PM_HOME = {"programGroups": [{"masterRecordId": "grp-1", "title": "STEM"}]}
_PM_GROUP = {"programs": [
    {"masterRecordId": "prog-1", "title": "Biology AS-T", "awardShortTitle": "AS-T"}]}
_PM_PROGRAM = {"pathways": [{"defaultPathway": True, "programMapId": "map-1"}]}


def _element(code, title, term):
    return {
        "name": code, "shortDescription": title,
        "recommendedOpportunity": {
            "type": "COURSE", "courseCode": code, "courseName": title,
            "minUnits": 4, "term": {"termNumber": term},
        },
    }


_PM_MAP = {"pathwayElements": [
    _element("ANATOMY 1", "Human Anatomy", 1),
    _element("BIOLOGY 3", "Introductory Biology", 1),
    _element("BIOLOGY 5", "General Biology", 2),
]}


def _routes():
    return {
        "/listing/LAMC/": _LISTING,
        "/home-page-content": _PM_HOME,
        "/program-groups/grp-1": _PM_GROUP,
        "/programs/prog-1": _PM_PROGRAM,
        "/program-maps/map-1": _PM_MAP,
        "/public/courses": json.loads(pathlib.Path(ELUMEN_FIXTURE).read_text()),
    }


def _catalog_prereqs(xlsx_path):
    """Return {Course ID -> Prerequisites (structured)} from the catalog sheet."""
    wb = openpyxl.load_workbook(xlsx_path)
    assert "catalog" in wb.sheetnames, wb.sheetnames
    ws = wb["catalog"]
    headers = [c.value for c in ws[1]]
    ci = headers.index("Course ID")
    pi = headers.index("Prerequisites (structured)")
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        out[row[ci]] = row[pi]
    return out


def _prereq_detector(report):
    for d in report["inert_detectors"]:
        if d["detector"] == "prerequisite_ordering":
            return d
    raise AssertionError("no prerequisite_ordering detector in report")


def _planned_courses(engine_results):
    """All courses placed in any cohort plan, across every program."""
    seen = set()
    for prog in engine_results["programs"].values():
        for cohort in prog["cohorts"].values():
            if cohort and cohort.get("plan"):
                for term_courses in cohort["plan"].values():
                    seen.update(term_courses)
    return seen


def test_elumen_live_attaches_real_prereq_and_engine_consumes(tmp_path):
    out_path = str(tmp_path / "live_LAMC.xlsx")
    client = FakeClient(_routes())

    report = build_live_workbook.analyze_live(
        campus="LAMC", terms=[2268], program_query="Biology",
        out_path=out_path, client=client, elumen_live=True,
    )

    assert report["error"] is None, report["error"]

    # The gated course's structured prereq is the golden CNF from the fixture.
    prereqs = _catalog_prereqs(out_path)
    assert prereqs.get("ANATOMY 1") == "(BIOLOGY 3 OR BIOLOGY 5)", prereqs

    # The coverage report is present (the eLumen<->catalog join audit).
    assert "elumen_coverage" in report, sorted(report.keys())
    cov = report["elumen_coverage"]
    assert "courses_fetched" in cov and "courses_with_prereqs" in cov, cov
    assert cov["courses_with_prereqs"] >= 1, cov

    # The prerequisite_ordering detector is ACTIVE, labelled REAL (not
    # fixture-only), carries live=True, and includes the coverage dict.
    det = _prereq_detector(report)
    assert det["status"] == "active", det
    assert det.get("live") is True, det
    assert "REAL eLumen" in det["label"], det["label"]
    assert "FIXTURE-ONLY" not in det["label"], det["label"]
    assert "fixture-only" not in det["label"].lower(), det["label"]
    assert det.get("coverage") == cov, det

    # engine.run consumed the finished workbook: it recognizes the program and
    # the solver places the gated course in a cohort plan.
    res = engine.run(out_path)
    assert isinstance(res, dict)
    assert res["programs"], res
    assert "ANATOMY 1" in _planned_courses(res), _planned_courses(res)

    # No warning when only --elumen-live is given.
    assert not report.get("warnings"), report.get("warnings")


def test_elumen_live_wins_over_fixture_with_warning(tmp_path):
    out_path = str(tmp_path / "live_wins.xlsx")
    client = FakeClient(_routes())

    report = build_live_workbook.analyze_live(
        campus="LAMC", terms=[2268], program_query="Biology",
        out_path=out_path, client=client,
        elumen_live=True, elumen_fixture=ELUMEN_FIXTURE,
    )

    assert report["error"] is None, report["error"]

    # A warning records the override (live wins, fixture ignored).
    assert report.get("warnings"), report
    joined = " ".join(report["warnings"])
    assert "elumen-live" in joined and "elumen-fixture" in joined, joined

    # The result is still the LIVE one: coverage present + detector labelled REAL.
    assert "elumen_coverage" in report, sorted(report.keys())
    det = _prereq_detector(report)
    assert det.get("live") is True, det
    assert "REAL eLumen" in det["label"], det["label"]

    # And the live prereq actually landed in the catalog (fixture would too, but
    # the warning + REAL label prove it took the LIVE branch).
    prereqs = _catalog_prereqs(out_path)
    assert prereqs.get("ANATOMY 1") == "(BIOLOGY 3 OR BIOLOGY 5)", prereqs


def test_no_network_socket_used(tmp_path, monkeypatch):
    """The whole build must run with ONLY the FakeClient: assert no real socket is
    created (the build path opens httpx.Client only when client is None, which
    never happens here)."""
    import socket

    def _boom(*a, **k):
        raise AssertionError("network socket attempted; the build is not offline")

    monkeypatch.setattr(socket.socket, "connect", _boom)

    out_path = str(tmp_path / "no_net.xlsx")
    client = FakeClient(_routes())
    report = build_live_workbook.analyze_live(
        campus="LAMC", terms=[2268], program_query="Biology",
        out_path=out_path, client=client, elumen_live=True,
    )
    assert report["error"] is None, report["error"]
    assert "elumen_coverage" in report, sorted(report.keys())
